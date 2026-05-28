"""
app.py — HAGAP_WEB v2.3
CORREÇÕES:
  → /api/baixar_db implementado (rodar_gerador.py precisava)
  → /api/migrar_db aceita dados JSON direto no body
  → api_dados(): auto-marca executado quando há ffos ou medicao_final
  → Suporte a múltiplas AEs por projeto (campo 'aes' lista)
  → BMDs sem projeto ficam como registros 'orfão' no DB
  → Projetos com FFO/medicao final → status Final automaticamente
"""

import json, os, re, base64, io
from datetime import date, datetime
from flask import Flask, jsonify, request, send_from_directory
import pandas as pd

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

@app.after_request
def _add_cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

DATABASE_URL  = os.environ.get("DATABASE_URL", "")
DB_PATH       = "db.json"
EXCEL_PATH    = "resultado.xlsx"
UPLOAD_FOLDER = "uploads"

if DATABASE_URL:
    PASTA_AES      = "pdfs/aes"
    PASTA_PROJETOS = "pdfs/projetos"
    PASTA_MEDICOES = "pdfs/medicoes"
else:
    PASTA_AES      = os.environ.get("PASTA_AES",      r"C:\HAGAP\AES")
    PASTA_PROJETOS = os.environ.get("PASTA_PROJETOS",  r"C:\HAGAP\PROJETOS")
    PASTA_MEDICOES = os.environ.get("PASTA_MEDICOES",  r"C:\HAGAP\medicoes")

STATUS_VALIDOS = {"Pendente", "Medição 1", "Medição 2", "Final", "Cancelado"}
CAMPOS_EXTRAS  = ("pde", "obs_parcial")  # campos livres sempre permitidos


# ─────────────────────────────────────────────────────────────────────
#  PostgreSQL
# ─────────────────────────────────────────────────────────────────────

def _pg_conn():
    import psycopg2
    url = DATABASE_URL
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    return psycopg2.connect(url)


def _pg_init():
    with _pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS hagap_db (
                    id   INT PRIMARY KEY DEFAULT 1,
                    dados JSONB NOT NULL DEFAULT '[]'
                )
            """)
            cur.execute("""
                INSERT INTO hagap_db (id, dados) VALUES (1, '[]')
                ON CONFLICT (id) DO NOTHING
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS hagap_mat_overrides (
                    id           SERIAL PRIMARY KEY,
                    projeto      TEXT NOT NULL,
                    arquivo      TEXT NOT NULL,
                    item_key     TEXT NOT NULL,
                    descricao    TEXT,
                    quantidade   FLOAT,
                    valor_unitario FLOAT,
                    valor_total  FLOAT,
                    is_deleted   BOOLEAN DEFAULT FALSE,
                    is_checked   BOOLEAN DEFAULT FALSE,
                    check_obs    TEXT DEFAULT '',
                    updated_at   TIMESTAMP DEFAULT NOW(),
                    UNIQUE (projeto, arquivo, item_key)
                )
            """)
        conn.commit()


_pg_ready = False
_pg_files_ready = False

_GOOGLE_SERVICE_ACCOUNT_B64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")


def _get_drive_service():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    json_b64 = _GOOGLE_SERVICE_ACCOUNT_B64
    if not json_b64:
        return None
    try:
        json_b64 += "=" * (-len(json_b64) % 4)
        info = json.loads(base64.b64decode(json_b64).decode("utf-8"))
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/drive"])
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        print(f"[Drive] Erro: {e}")
        return None


def _find_drive_folder(service, folder_name="hagap"):
    if not service:
        return None
    return os.environ.get("DRIVE_FOLDER_ID", "1yWPVwMMmSfDItku95Bq6lLGlhSkLrS0u")


def _pg_files_init():
    with _pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS hagap_files (
                    pasta   TEXT NOT NULL,
                    nome    TEXT NOT NULL,
                    conteudo BYTEA NOT NULL,
                    mimetype TEXT DEFAULT 'application/octet-stream',
                    criado_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (pasta, nome)
                )
            """)
        conn.commit()


def _salvar_arquivo_pg(pasta: str, nome: str, conteudo: bytes, mimetype: str = 'application/octet-stream'):
    """Salva ou substitui um arquivo no PostgreSQL."""
    try:
        _ensure_pg_files()
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hagap_files (pasta, nome, conteudo, mimetype)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (pasta, nome) DO UPDATE
                        SET conteudo = EXCLUDED.conteudo,
                            mimetype = EXCLUDED.mimetype,
                            criado_at = NOW()
                """, (pasta, nome, conteudo, mimetype))
            conn.commit()
        return True
    except Exception as e:
        print(f'[Files] Erro ao salvar {pasta}/{nome}: {e}')
        return False


def _ler_arquivo_pg(pasta: str, nome: str):
    """Retorna (conteudo_bytes, mimetype) ou (None, None) se não encontrado."""
    try:
        _ensure_pg_files()
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT conteudo, mimetype FROM hagap_files WHERE pasta=%s AND nome=%s',
                    (pasta, nome)
                )
                row = cur.fetchone()
                if row:
                    conteudo = bytes(row[0]) if not isinstance(row[0], bytes) else row[0]
                    return conteudo, row[1]
        return None, None
    except Exception as e:
        print(f'[Files] Erro ao ler {pasta}/{nome}: {e}')
        return None, None


def _listar_arquivos_pg(pasta: str):
    """Lista nomes de arquivos de uma pasta no PostgreSQL."""
    try:
        _ensure_pg_files()
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT nome FROM hagap_files WHERE pasta=%s ORDER BY nome',
                    (pasta,)
                )
                return [row[0] for row in cur.fetchall()]
    except Exception as e:
        print(f'[Files] Erro ao listar {pasta}: {e}')
        return []


def _ensure_pg_files():
    global _pg_files_ready
    if not _pg_files_ready:
        _ensure_pg()  # garante que a conexão PG está ok
        _pg_files_init()
        _pg_files_ready = True


def _ensure_pg():
    global _pg_ready
    if not _pg_ready:
        _pg_init()
        _pg_ready = True
        _auto_migrar_db()


def _auto_migrar_db():
    if not os.path.exists(DB_PATH):
        return
    try:
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT dados FROM hagap_db WHERE id=1")
                row = cur.fetchone()
                dados_pg = row[0] if row else []
                if isinstance(dados_pg, str):
                    dados_pg = json.loads(dados_pg)
                if dados_pg:
                    return
        with open(DB_PATH, encoding="utf-8") as f:
            dados = json.load(f)
        if not isinstance(dados, list) or not dados:
            return
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hagap_db (id, dados) VALUES (1, %s)
                    ON CONFLICT (id) DO UPDATE SET dados = EXCLUDED.dados
                """, (json.dumps(dados, ensure_ascii=False),))
            conn.commit()
        print(f"[DB] ✅ Migração automática: {len(dados)} registros")
    except Exception as e:
        print(f"[DB] ⚠️ Erro na migração: {e}")


# ─────────────────────────────────────────────────────────────────────
#  DB helpers
# ─────────────────────────────────────────────────────────────────────

def _auto_corrigir(dados: list) -> list:
    """Garante consistência: projetos com ffos/final → executado=True, status=Final"""
    for d in dados:
        # Normaliza nome do projeto (remove espaços, \r, \n, \t)
        if isinstance(d.get('projeto'), str):
            d['projeto'] = re.sub(r'[\r\n\t]', '', d['projeto'].strip())
        if d.get("ffos") or d.get("final"):
            d["executado"] = True
            if d.get("status") not in ("Final", "Cancelado"):
                d["status"] = "Final"
        # us_num sempre numérico
        if not d.get("us_num") and d.get("us"):
            d["us_num"] = _to_float(d["us"])
        # garantir campos lista
        for campo in ("bmds", "ffos", "pdfs_projeto", "avisos", "origem"):
            if not isinstance(d.get(campo), list):
                d[campo] = []
    return dados


def carregar_db_pg() -> list:
    """Lê os dados atuais do PostgreSQL (sem fallback para disco)."""
    _ensure_pg()
    with _pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT dados FROM hagap_db WHERE id=1")
            row = cur.fetchone()
            if row and row[0]:
                d = row[0]
                return d if isinstance(d, list) else json.loads(d)
    return []


def ler_db() -> list:
    if DATABASE_URL:
        try:
            _ensure_pg()
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT dados FROM hagap_db WHERE id=1")
                    row = cur.fetchone()
                    if row and row[0]:
                        d = row[0]
                        dados = d if isinstance(d, list) else json.loads(d)
                        return _auto_corrigir(dados)
            return []
        except Exception as e:
            print(f"[DB] Erro PostgreSQL: {e}")
            return []
    if not os.path.exists(DB_PATH):
        return []
    with open(DB_PATH, encoding="utf-8") as f:
        conteudo = f.read().strip()
    if not conteudo:
        return []
    try:
        dados = json.loads(conteudo)
        if isinstance(dados, list):
            return _auto_corrigir(dados)
        if isinstance(dados, dict):
            return _auto_corrigir([dados])
    except json.JSONDecodeError:
        pass
    return []


def salvar_db(dados: list):
    dados = _auto_corrigir(dados)
    if DATABASE_URL:
        try:
            _ensure_pg()
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        INSERT INTO hagap_db (id, dados) VALUES (1, %s)
                        ON CONFLICT (id) DO UPDATE SET dados = EXCLUDED.dados
                    """, (json.dumps(dados, ensure_ascii=False),))
                conn.commit()
            return
        except Exception as e:
            print(f"[DB] Erro ao salvar: {e}")
            return
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def _norm_proj(s: str) -> str:
    """Normaliza nome de projeto para comparação: strip, uppercase, remove \r\n\t."""
    return re.sub(r'[\r\n\t]', '', (s or '').strip()).upper()


def encontrar_idx(dados: list, projeto: str) -> int:
    projeto_norm = _norm_proj(projeto)
    # Passo 1: comparação exata (mais rápida)
    for i, d in enumerate(dados):
        if d.get('projeto', '') == projeto:
            return i
    # Passo 2: comparação normalizada (sem espaços, case-insensitive, sem \r)
    for i, d in enumerate(dados):
        if _norm_proj(d.get('projeto', '')) == projeto_norm:
            return i
    return -1


# ─────────────────────────────────────────────────────────────────────
#  Utilitários
# ─────────────────────────────────────────────────────────────────────

def _to_float(s):
    if not s:
        return 0.0
    s = re.sub(r"[^\d,\.]", "", str(s))
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fmt_brl(v):
    if v is None:
        return "—"
    try:
        s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {s}" if v >= 0 else f"-R$ {s}"
    except Exception:
        return "—"


def _parsear_data(s):
    if not s:
        return None
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"]
    s_clean = re.sub(r"\s+", " ", s.strip().split()[0])
    for fmt in formatos:
        try:
            return datetime.strptime(s_clean, fmt).date()
        except ValueError:
            pass
    return None


def _alerta_prazo(prazo_str):
    d = _parsear_data(prazo_str)
    if not d:
        return ""
    diff = (d - date.today()).days
    if diff < 0:   return "vencido"
    if diff <= 7:  return "critico"
    if diff <= 30: return "alerta"
    return "ok"


# ─────────────────────────────────────────────────────────────────────
#  Extração de PDF
# ─────────────────────────────────────────────────────────────────────

def _extrair_texto_medicao(caminho):
    import pdfplumber
    texto = ""
    try:
        with pdfplumber.open(caminho) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    texto += t + "\n"
    except Exception:
        pass
    if not texto.strip():
        try:
            import fitz
            doc = fitz.open(caminho)
            for p in doc:
                texto += p.get_text()
        except Exception:
            pass
    return texto


def _extrair_rateio_e_us(texto):
    rateio_str = ""
    us_str = ""
    m = re.search(
        r'RATEIO.{0,60}CUSTO.{0,60}TOTAL.{0,200}HAGAP.{0,300}R\$\s*([\d\.]+,\d{2})',
        texto, re.IGNORECASE | re.DOTALL
    )
    if m:
        rateio_str = m.group(1).strip()
    if not rateio_str:
        for linha in texto.split("\n"):
            lu = linha.upper().replace(' ', '')
            if 'RATEIO' in lu and 'HAGAP' in lu:
                mv = re.search(r'R\$\s*([\d\.]+,\d{2})', linha)
                if mv:
                    rateio_str = mv.group(1).strip()
                    break
    if not rateio_str:
        for linha in texto.split("\n"):
            lu = linha.upper().replace(' ', '')
            if 'RATEIO' in lu and 'CUSTO' in lu:
                mv = re.search(r'R\$\s*([\d\.]+,\d{2})', linha)
                if mv:
                    rateio_str = mv.group(1).strip()
                    break
    if not rateio_str:
        for linha in texto.split("\n"):
            if 'RATEIO' in linha.upper():
                mv = re.search(r'R\$\s*([\d\.]+,\d{2})', linha)
                if mv:
                    rateio_str = mv.group(1).strip()
                    break
    m_us = re.search(r'TOTALM\.O\.\s+([\d,]+)', texto, re.IGNORECASE)
    if m_us:
        us_str = m_us.group(1)
    return rateio_str, us_str


# ─────────────────────────────────────────────────────────────────────
#  Rotas estáticas
# ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    from flask import make_response

    resp = make_response(send_from_directory("templates", "index.html"))

    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"

    return resp
# ─── BOLETIM DE OBRAS (BDO) — acesso público, sem login ───
@app.route("/boletim")
def boletim():
    return send_from_directory("templates", "boletim.html")
# ─────────────────────────────────────────────────────────────────────
#  API: dados
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/dados")
def api_dados():
    try:
        dados = ler_db()
    except Exception as e:
        return jsonify({"erro": f"Falha ao ler db: {str(e)}"}), 500
    for d in dados:
        d["prazo_alerta"] = _alerta_prazo(d.get("prazo", ""))
    return jsonify(dados)


# ─────────────────────────────────────────────────────────────────────
#  API: baixar DB (para rodar_gerador.py)
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/baixar_db")
def api_baixar_db():
    SENHA_MIGRACAO = os.environ.get("SENHA_MIGRACAO", "HAGAP_MIGRAR")
    senha = request.args.get("senha", "")
    if senha != SENHA_MIGRACAO:
        return jsonify({"erro": "Senha incorreta"}), 403
    dados = ler_db()
    return jsonify(dados)


# ─────────────────────────────────────────────────────────────────────
#  API: migrar / sincronizar DB
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/migrar_db", methods=["GET", "POST"])
def api_migrar_db():
    SENHA_MIGRACAO = os.environ.get("SENHA_MIGRACAO", "HAGAP_MIGRAR")
    senha = request.args.get("senha", "") or (request.get_json(silent=True) or {}).get("senha", "")
    if senha != SENHA_MIGRACAO:
        return jsonify({"erro": "Senha incorreta"}), 403
    if not DATABASE_URL:
        return jsonify({"erro": "DATABASE_URL não configurado"}), 400

    # Campos editados manualmente pelo usuário — NUNCA sobrescrever com dados do gerador
    CAMPOS_MANUAIS_FIXOS = (
        'pde', 'pde_data', 'comentario_parcial', 'obs_parcial',
        'pasta_projeto_url',  # URL do Drive corrigida manualmente
    )

    def _mesclar_com_pg(novos: list) -> list:
        """Mescla novos registros do gerador com os dados existentes no PG,
        preservando TODOS os campos editados manualmente pelo usuário."""
        try:
            dados_pg = carregar_db_pg()
        except Exception:
            dados_pg = []
        idx_pg = {re.sub(r'[\s\-_]', '', (r.get('projeto') or '')).upper(): r
                  for r in dados_pg}
        resultado = []
        for novo in novos:
            k = re.sub(r'[\s\-_]', '', (novo.get('projeto') or '')).upper()
            existente = idx_pg.get(k)
            if existente:
                merged = dict(novo)  # começa com dados frescos do gerador

                # 1. Campos fixos manuais (PDE, comentário etc.) — sempre preserva
                for campo in CAMPOS_MANUAIS_FIXOS:
                    if existente.get(campo) not in (None, '', False, 0):
                        merged[campo] = existente[campo]

                # 2. Campos rastreados como editados manualmente via /api/editar
                #    — só o usuário pode sobrescrever, gerador nunca apaga
                campos_manuais_usuario = existente.get('_campos_manuais', [])
                for campo in campos_manuais_usuario:
                    val = existente.get(campo)
                    if val not in (None, '', False, 0, []):
                        merged[campo] = val
                merged['_campos_manuais'] = campos_manuais_usuario

                # 3. Vínculos de PDF: MERGE das listas (nunca descarta links manuais)
                #    Garante que PDFs vinculados manualmente não sejam perdidos
                aes_novo   = novo.get('aes', [])
                aes_pg     = existente.get('aes', [])
                arqs_novo  = {a.get('arquivo') for a in aes_novo}
                for ae_item in aes_pg:
                    if ae_item.get('arquivo') not in arqs_novo:
                        aes_novo.append(ae_item)  # preserva link manual
                merged['aes'] = aes_novo
                # arquivo_ae (compat legado): preserva se existente e gerador não tem
                if existente.get('arquivo_ae') and not novo.get('arquivo_ae'):
                    merged['arquivo_ae'] = existente['arquivo_ae']

                # 4. Status/execução manual
                status_pg = existente.get('status')
                if status_pg == 'Cancelado':
                    merged['status'] = 'Cancelado'
                    merged['executado'] = False
                elif status_pg in ('Medição 1', 'Medição 2') and not novo.get('executado'):
                    merged['status'] = status_pg
                    merged['executado'] = False
                    for mc in ('medicao1', 'medicao2'):
                        if existente.get(mc) and not novo.get(mc):
                            merged[mc] = existente[mc]
                elif existente.get('executado') and not novo.get('executado'):
                    merged['executado'] = True
                    if status_pg == 'Final':
                        merged['status'] = 'Final'

                resultado.append(merged)
                idx_pg.pop(k, None)
            else:
                resultado.append(novo)
        # Adiciona registros criados manualmente no site (não existem no gerador)
        for k_pg, r_pg in idx_pg.items():
            if r_pg.get('projeto', '').startswith('ORFAO_'):
                continue
            resultado.append(r_pg)
        return resultado

    # Prioridade 1: dados enviados diretamente no body (rodar_gerador.py)
    body = request.get_json(force=True, silent=True) or {}
    dados_body = body.get("dados")
    if dados_body and isinstance(dados_body, list):
        dados_mesclados = _mesclar_com_pg(dados_body)
        salvar_db(dados_mesclados)
        return jsonify({"ok": True, "registros": len(dados_mesclados),
                        "novos": len(dados_body), "fonte": "body",
                        "preservados": len(dados_mesclados) - len(dados_body)})

    # Prioridade 2: db.json local no servidor (legado)
    if not os.path.exists(DB_PATH):
        return jsonify({"erro": "db.json não encontrado e nenhum dado no body"}), 404
    with open(DB_PATH, encoding="utf-8") as f:
        dados = json.load(f)
    dados_mesclados = _mesclar_com_pg(dados)
    salvar_db(dados_mesclados)
    return jsonify({"ok": True, "registros": len(dados_mesclados), "fonte": "db.json"})


@app.route("/api/health")
def api_health():
    db_existe  = os.path.exists(DB_PATH)
    db_tamanho = os.path.getsize(DB_PATH) if db_existe else 0
    qtd = 0
    erro_db = None
    if db_existe:
        try:
            qtd = len(ler_db())
        except Exception as e:
            erro_db = str(e)
    return jsonify({
        "status":    "ok" if db_existe and not erro_db else "erro",
        "db_existe": db_existe, "db_bytes": db_tamanho,
        "registros": qtd, "erro_db": erro_db,
        "pastas": {
            "aes":      os.path.isdir(PASTA_AES),
            "projetos": os.path.isdir(PASTA_PROJETOS),
            "medicoes": os.path.isdir(PASTA_MEDICOES),
        }
    })


# ─────────────────────────────────────────────────────────────────────
#  API: editar
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/editar", methods=["POST"])
def api_editar():
    body    = request.get_json() or {}
    projeto = (body.get("projeto_original") or body.get("projeto") or "").strip()
    campos  = body.get("campos") or {k: v for k, v in body.items() if k != "projeto_original"}
    if not projeto or not campos:
        return jsonify({"erro": "Dados inválidos"}), 400
    dados = ler_db()
    idx = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    status_novo = campos.get("status")
    if status_novo and status_novo not in STATUS_VALIDOS:
        return jsonify({"erro": f"Status inválido: {status_novo}"}), 400
    # Rastreia quais campos foram editados manualmente (para o merge com o gerador)
    campos_manuais_set = set(dados[idx].get('_campos_manuais', []))
    for campo, valor in campos.items():
        dados[idx][campo] = valor
        if not campo.startswith('_'):
            campos_manuais_set.add(campo)
    dados[idx]['_campos_manuais'] = list(campos_manuais_set)
    if "us" in campos:
        dados[idx]["us_num"] = _to_float(campos["us"])
    if "prazo" in campos:
        dados[idx]["prazo_alerta"] = _alerta_prazo(campos["prazo"])
    # Auto-executado
    if dados[idx].get("status") == "Final" or dados[idx].get("executado"):
        dados[idx]["executado"] = True
        dados[idx]["status"] = "Final"
    elif dados[idx].get("ffos") or dados[idx].get("final"):
        dados[idx]["executado"] = True
        dados[idx]["status"] = "Final"
    elif dados[idx].get("status") in ("Pendente", "Medição 1", "Medição 2"):
        dados[idx]["executado"] = False
    salvar_db(dados)
    return jsonify({"ok": True, "registro": dados[idx]})


# ─────────────────────────────────────────────────────────────────────
#  API: adicionar
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/adicionar", methods=["POST"])
def api_adicionar():
    body    = request.get_json() or {}
    projeto = (body.get("projeto") or "").strip()
    if not projeto:
        return jsonify({"erro": "Campo 'projeto' obrigatório"}), 400
    dados = ler_db()
    if encontrar_idx(dados, projeto) != -1:
        return jsonify({"erro": "Projeto já existe"}), 409
    novo = {
        "projeto":       projeto,
        "ae":            body.get("ae", ""),
        "aes":           body.get("aes", []),  # lista de AEs
        "us":            body.get("us", ""),
        "us_num":        _to_float(body.get("us", "")),
        "prazo":         body.get("prazo", ""),
        "prazo_alerta":  _alerta_prazo(body.get("prazo", "")),
        "local":         body.get("local", ""),
        "custo":         body.get("custo", ""),
        "status":        body.get("status", "Pendente"),
        "executado":     body.get("executado", False),
        "arquivo_ae":    body.get("arquivo_ae", ""),
        "pasta_projeto": body.get("pasta_projeto", ""),
        "pdfs_projeto":  [],
        "bmds":          [],
        "ffos":          [],
        "medicao1":      body.get("medicao1", ""),
        "medicao2":      body.get("medicao2", ""),
        "final":         body.get("final", ""),
        "origem":        ["manual"],
        "avisos":        [],
    }
    dados.append(novo)
    salvar_db(dados)
    return jsonify({"ok": True, "registro": novo})


# ─────────────────────────────────────────────────────────────────────
#  API: excluir
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/excluir", methods=["POST"])
def api_excluir():
    body    = request.get_json() or {}
    projeto = body.get("projeto", "").strip()
    if not projeto:
        return jsonify({"erro": "Campo 'projeto' obrigatório"}), 400
    dados = ler_db()
    idx = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    dados.pop(idx)
    salvar_db(dados)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────
#  API: linkar AE / BMD / FFO
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/linkar_multiplos", methods=["POST"])
def api_linkar_multiplos():
    """Vincula múltiplos arquivos a um projeto em uma única chamada."""
    body    = request.get_json() or {}
    projeto = body.get("projeto", "").strip()
    arquivos = body.get("arquivos", [])  # lista de {arquivo, tipo, url, data, us, valor}
    if not projeto or not arquivos:
        return jsonify({"erro": "Campos 'projeto' e 'arquivos' são obrigatórios"}), 400
    dados = ler_db()
    idx = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    resultados = []
    for item in arquivos:
        arquivo = item.get("arquivo", "").strip()
        tipo    = item.get("tipo", "outro").lower()
        if not arquivo:
            continue
        extras = {"url": item.get("url", ""), "data": item.get("data", ""),
                  "us": item.get("us", ""), "valor": item.get("valor", ""),
                  "ae": item.get("ae", "")}
        r = dados[idx]
        if tipo == "ae":
            r["arquivo_ae"] = r["arquivo_ae"] or arquivo
            r.setdefault("aes", [])
            if arquivo not in [a.get("arquivo") for a in r["aes"]]:
                r["aes"].append({"arquivo": arquivo, "url": extras.get("url", ""), "ae": extras.get("ae", "")})
        elif tipo == "bmd":
            entrada = {"arquivo": arquivo, **{k: v for k, v in extras.items() if v}}
            if arquivo not in [b.get("arquivo") for b in r.get("bmds", [])]:
                r.setdefault("bmds", []).append(entrada)
                if not r.get("medicao1"):
                    r["medicao1"] = extras.get("data") or arquivo
                elif not r.get("medicao2"):
                    r["medicao2"] = extras.get("data") or arquivo
        elif tipo == "ffo":
            entrada = {"arquivo": arquivo, **{k: v for k, v in extras.items() if v}}
            if arquivo not in [f.get("arquivo") for f in r.get("ffos", [])]:
                r.setdefault("ffos", []).append(entrada)
                if not r.get("final"):
                    r["final"] = extras.get("data") or arquivo
                r["executado"] = True
                r["status"]    = "Final"
        else:
            entrada = {"arquivo": arquivo, "tipo": tipo, **{k: v for k, v in extras.items() if v}}
            r.setdefault("arquivos_extras", [])
            if not any(a.get("arquivo") == arquivo for a in r["arquivos_extras"]):
                r["arquivos_extras"].append(entrada)
        resultados.append({"arquivo": arquivo, "tipo": tipo, "ok": True})
        if "manual" not in r.get("origem", []):
            r.setdefault("origem", []).append("manual")
        # Rastreia vínculos como editados manualmente
        campos_manuais_set = set(r.get('_campos_manuais', []))
        for campo_vinculo in ('aes', 'arquivo_ae', 'bmds', 'ffos', 'arquivos_extras', 'medicao1', 'medicao2', 'final', 'executado', 'status'):
            campos_manuais_set.add(campo_vinculo)
        r['_campos_manuais'] = list(campos_manuais_set)
    salvar_db(dados)
    return jsonify({"ok": True, "vinculados": len(resultados), "detalhes": resultados, "registro": dados[idx]})


@app.route("/api/linkar", methods=["POST"])
def api_linkar():
    body    = request.get_json() or {}
    projeto = body.get("projeto", "").strip()
    tipo    = body.get("tipo", "").lower()
    arquivo = body.get("arquivo", "").strip()
    if not projeto or not tipo or not arquivo:
        return jsonify({"erro": "Campos 'projeto', 'tipo' e 'arquivo' são obrigatórios"}), 400
    TIPOS_VALIDOS = {"ae", "bmd", "ffo", "art", "contrato", "projeto", "orcamento", "nf", "outro"}
    if tipo not in TIPOS_VALIDOS:
        return jsonify({"erro": f"Tipo inválido: {tipo}"}), 400
    dados = ler_db()
    idx = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    r      = dados[idx]
    extras = body.get("dados_extras", {})
    if tipo == "ae":
        r["arquivo_ae"] = arquivo
        # adicionar à lista de AEs também
        r.setdefault("aes", [])
        nova_ae = {"arquivo": arquivo, "url": extras.get("url", ""), "ae": extras.get("ae", "")}
        if arquivo not in [a.get("arquivo") for a in r["aes"]]:
            r["aes"].append(nova_ae)
        for k in ("ae", "us", "prazo", "local", "custo"):
            if k in extras and not r.get(k):
                r[k] = extras[k]
    elif tipo == "bmd":
        entrada = {"arquivo": arquivo, **extras}
        if "rateio" not in extras:
            cam = os.path.join(PASTA_MEDICOES, arquivo)
            if os.path.exists(cam):
                texto_med = _extrair_texto_medicao(cam)
                rateio_str, us_med = _extrair_rateio_e_us(texto_med)
                if rateio_str:
                    entrada["rateio"]       = rateio_str
                    entrada["rateio_float"] = _to_float(rateio_str)
                if us_med and not entrada.get("us"):
                    entrada["us"] = us_med
        if arquivo not in [b.get("arquivo") for b in r.get("bmds", [])]:
            r.setdefault("bmds", []).append(entrada)
            if not r.get("medicao1"):
                r["medicao1"] = extras.get("data") or arquivo
            elif not r.get("medicao2"):
                r["medicao2"] = extras.get("data") or arquivo
    elif tipo == "ffo":
        entrada = {"arquivo": arquivo, **extras}
        if arquivo not in [f.get("arquivo") for f in r.get("ffos", [])]:
            r.setdefault("ffos", []).append(entrada)
            if not r.get("final"):
                r["final"] = extras.get("data") or arquivo
            r["executado"] = True
            r["status"]    = "Final"
    else:
        entrada = {"arquivo": arquivo, "tipo": tipo, **extras}
        r.setdefault("arquivos_extras", [])
        if not any(a.get("arquivo") == arquivo for a in r["arquivos_extras"]):
            r["arquivos_extras"].append(entrada)
    if "manual" not in r.get("origem", []):
        r.setdefault("origem", []).append("manual")
    # Rastreia vínculos como editados manualmente (protege contra overwrite do gerador)
    _cm = set(r.get('_campos_manuais', []))
    for _cv in ('aes','arquivo_ae','bmds','ffos','arquivos_extras','medicao1','medicao2','final','executado','status'):
        _cm.add(_cv)
    r['_campos_manuais'] = list(_cm)
    salvar_db(dados)
    return jsonify({"ok": True, "registro": r})


# ─────────────────────────────────────────────────────────────────────
#  API: arquivos do projeto
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/arquivos_projeto")
def api_arquivos_projeto():
    projeto = request.args.get("projeto", "").strip()
    dados   = ler_db()
    idx     = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"pdfs": [], "extras": []})
    r         = dados[idx]
    pasta_url = r.get("pasta_projeto_url", "")
    pdfs      = []
    # Múltiplas AEs
    aes_lista = r.get("aes", [])
    if aes_lista:
        for ae_item in aes_lista:
            arq = ae_item.get("arquivo", "")
            if arq:
                pdfs.append({"arquivo": arq, "url_drive": ae_item.get("url", ""),
                             "url_local": f"/pdfs/aes/{arq}", "tipo": "ae"})
    elif r.get("arquivo_ae"):
        pdfs.append({"arquivo": r["arquivo_ae"], "url_drive": r.get("arquivo_ae_url", ""),
                     "url_local": f"/pdfs/aes/{r['arquivo_ae']}", "tipo": "ae"})
    for b in r.get("bmds", []):
        arq = b.get("arquivo", "")
        if arq:
            pdfs.append({"arquivo": arq, "url_drive": b.get("url", ""),
                         "url_local": f"/pdfs/medicoes/{arq}", "tipo": "bmd",
                         "rateio": b.get("rateio", ""), "rateio_float": b.get("rateio_float", 0),
                         "data": b.get("data", "")})
    for f in r.get("ffos", []):
        arq = f.get("arquivo", "")
        if arq:
            pdfs.append({"arquivo": arq, "url_drive": f.get("url", ""),
                         "url_local": f"/pdfs/medicoes/{arq}", "tipo": "ffo",
                         "rateio": f.get("rateio", ""), "rateio_float": f.get("rateio_float", 0),
                         "data": f.get("data", "")})
    extras = r.get("arquivos_extras", [])
    for ex in extras:
        arq   = ex.get("arquivo", "")
        # Usa pasta_projeto ou, como fallback, o número do projeto (para arquivos enviados via modal vincular)
        pasta = r.get("pasta_projeto", "") or r.get("projeto", "")
        if arq:
            pdfs.append({"arquivo": arq, "url_drive": ex.get("url", ""),
                         "url_local": f"/pdfs/projetos/{pasta}/{arq}" if pasta else f"/uploads/{arq}",
                         "tipo": ex.get("tipo", "outro")})
    return jsonify({"pdfs": pdfs, "extras": extras, "pasta_projeto_url": pasta_url})


# ─────────────────────────────────────────────────────────────────────
#  API: somar US + financeiro (lucro/prejuízo)
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/soma_us")
def api_soma_us():
    mes         = request.args.get("mes",         "").strip()
    ano         = request.args.get("ano",         "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim    = request.args.get("data_fim",    "").strip()
    todos_flag  = request.args.get("todos", "false").lower() in ("true", "1", "yes")
    dados       = ler_db()

    # Pré-processa filtros de data
    dt_ini = _parsear_data(data_inicio) if data_inicio else None
    dt_fim = _parsear_data(data_fim)    if data_fim    else None
    if data_inicio and not dt_ini:
        return jsonify({"erro": f"data_inicio inválida: '{data_inicio}'"}), 400
    if data_fim and not dt_fim:
        return jsonify({"erro": f"data_fim inválida: '{data_fim}'"}), 400

    tem_filtro_data = bool(mes or ano or data_inicio or data_fim)

    def _medicao_no_periodo(dt_str):
        """Verifica se a data de uma medição está dentro do período filtrado."""
        if not dt_str:
            return not tem_filtro_data  # sem data: inclui só se não há filtro
        dt = _parsear_data(dt_str)
        if not dt:
            return not tem_filtro_data
        if mes and ano:
            return str(dt.month).zfill(2) == mes.zfill(2) and str(dt.year) == ano
        if mes:
            return str(dt.month).zfill(2) == mes.zfill(2)
        if ano:
            return str(dt.year) == ano
        if dt_ini and dt_fim:
            return dt_ini <= dt <= dt_fim
        if dt_ini:
            return dt >= dt_ini
        if dt_fim:
            return dt <= dt_fim
        return True

    total_us       = 0.0
    total_rateio   = 0.0
    total_custo_ae = 0.0
    projetos       = []

    for d in dados:
        bmds_todos = d.get("bmds", [])
        ffos_todos = d.get("ffos", [])

        if tem_filtro_data:
            # Filtra cada BMD/FFO pela SUA PRÓPRIA data
            bmds_uso = [b for b in bmds_todos if _medicao_no_periodo(b.get("data", ""))]
            ffos_uso = [f for f in ffos_todos if _medicao_no_periodo(f.get("data", ""))]
            # Se nenhuma medição cai no período, pula o projeto
            if not bmds_uso and not ffos_uso:
                continue
        else:
            # Sem filtro de data: inclui todos executados (ou todos se todos_flag)
            eh_exec = (d.get("executado") or d.get("status") == "Final"
                       or d.get("ffos") or d.get("final"))
            tem_valor = any((b.get("rateio_float") or 0) > 0 for b in bmds_todos)
            if not todos_flag and not eh_exec:
                continue
            if todos_flag and not eh_exec and not tem_valor:
                continue
            bmds_uso = bmds_todos
            ffos_uso = ffos_todos

        # Soma rateio das medições filtradas
        medicoes_list = []
        proj_rateio   = 0.0
        for bmd in bmds_uso:
            rv = bmd.get("rateio_float") or _to_float(bmd.get("rateio", ""))
            proj_rateio += rv
            if bmd.get("arquivo"):
                medicoes_list.append({"arquivo": bmd["arquivo"], "tipo": "BMD",
                                      "data": bmd.get("data", ""),
                                      "rateio_fmt": _fmt_brl(rv) if rv else "—"})
        for ffo in ffos_uso:
            rv = ffo.get("rateio_float") or _to_float(ffo.get("rateio", ""))
            proj_rateio += rv
            if ffo.get("arquivo"):
                medicoes_list.append({"arquivo": ffo["arquivo"], "tipo": "FFO",
                                      "data": ffo.get("data", ""),
                                      "rateio_fmt": _fmt_brl(rv) if rv else "—"})

        custo_ae = _to_float(d.get("custo", ""))
        resultado_proj = proj_rateio - custo_ae
        pct_proj       = round(resultado_proj / custo_ae * 100, 2) if custo_ae > 0 else None
        situacao_proj  = "lucro" if resultado_proj > 0 else ("prejuizo" if resultado_proj < 0 else "")
        total_rateio   += proj_rateio
        total_custo_ae += custo_ae

        # Data de referência: último FFO ou último BMD filtrado
        data_exec = ""
        if ffos_uso:
            data_exec = sorted(ffos_uso, key=lambda x: x.get("data", "") or "")[-1].get("data", "")
        elif bmds_uso:
            data_exec = sorted(bmds_uso, key=lambda x: x.get("data", "") or "")[-1].get("data", "")

        us_proj = d.get("us_num", 0) or _to_float(d.get("us", ""))
        total_us += us_proj

        projetos.append({
            "projeto":       d["projeto"],
            "us":            d.get("us", ""),
            "us_num":        us_proj,
            "final":         d.get("final", ""),
            "data_exec":     data_exec,
            "rateio_fmt":    _fmt_brl(proj_rateio) if proj_rateio else "—",
            "custo_ae_fmt":  _fmt_brl(custo_ae) if custo_ae else "—",
            "resultado_fmt": _fmt_brl(resultado_proj),
            "situacao":      situacao_proj,
            "pct":           pct_proj,
            "medicoes":      medicoes_list,
        })

    resultado_geral = total_rateio - total_custo_ae
    pct_geral       = round(resultado_geral / total_custo_ae * 100, 2) if total_custo_ae > 0 else None
    situacao_geral  = "lucro" if resultado_geral > 0 else ("prejuizo" if resultado_geral < 0 else "")

    return jsonify({
        "total_us":            total_us,
        "qtd_projetos":        len(projetos),
        "projetos":            projetos,
        "filtro_mes":          mes,
        "filtro_ano":          ano,
        "filtro_inicio":       data_inicio,
        "filtro_fim":          data_fim,
        "total_rateio":        total_rateio,
        "total_rateio_fmt":    _fmt_brl(total_rateio),
        "total_custo_ae":      total_custo_ae,
        "total_custo_ae_fmt":  _fmt_brl(total_custo_ae),
        "resultado_geral":     resultado_geral,
        "resultado_geral_fmt": _fmt_brl(resultado_geral),
        "situacao_geral":      situacao_geral,
        "pct_geral":           pct_geral,
    })


# ─────────────────────────────────────────────────────────────────────
#  API: importar medicoes.xlsx → atualiza rateio no db
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/importar_medicoes", methods=["POST"])
def api_importar_medicoes():
    body         = request.get_json(force=True) if request.data else {}
    caminho_xlsx = body.get("arquivo", "medicoes.xlsx") or "medicoes.xlsx"

    if not os.path.exists(caminho_xlsx):
        caminho_xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)), caminho_xlsx)
    if not os.path.exists(caminho_xlsx):
        return jsonify({"erro": f"Arquivo não encontrado: {caminho_xlsx}"}), 404

    try:
        df = pd.read_excel(caminho_xlsx)
        df.columns = [c.strip() for c in df.columns]
    except Exception as e:
        return jsonify({"erro": f"Erro ao ler Excel: {e}"}), 500

    col_proj   = next((c for c in df.columns if c.lower() == "projeto"), None)
    col_valor  = next((c for c in df.columns if c.lower() in ("valor r$", "valor")), None)
    col_tipo   = next((c for c in df.columns if c.lower() == "tipo"), None)
    col_arq    = next((c for c in df.columns if c.lower() in ("arquivo", "arquivo pdf", "nome arquivo")), None)
    col_data   = next((c for c in df.columns if c.lower() in ("data", "data medicao", "data medi\u00e7\u00e3o", "competencia", "compet\u00eancia", "data de medi\u00e7\u00e3o", "data emiss\u00e3o")), None)

    if not col_proj or not col_valor:
        return jsonify({"erro": f"Colunas não encontradas. Encontradas: {list(df.columns)}"}), 400

    # Processa linha a linha (preserva Tipo e Arquivo individuais)
    linhas_por_proj: dict = {}  # proj_num -> list of {tipo, arquivo, valor, data}
    for _, row in df.iterrows():
        proj  = str(row[col_proj]).strip().upper()
        if not proj or proj in ("", "NAN", "NONE"):
            continue
        valor = _to_float(str(row.get(col_valor, 0) if col_valor else 0))
        tipo  = str(row.get(col_tipo, "BMD") if col_tipo else "BMD").strip().upper()
        if tipo not in ("BMD", "FFO", "FINAL", "PARCIAL"):
            tipo = "BMD"
        if tipo in ("FINAL", "PARCIAL"):
            tipo = "FFO" if tipo == "FINAL" else "BMD"
        arq  = str(row.get(col_arq, "") if col_arq else "").strip()
        if not arq or arq in ("NAN", "NONE", ""):
            arq = f"{proj}_{tipo}_importado.pdf"
        data = str(row.get(col_data, "") if col_data else "").strip()
        if data in ("NAN", "NONE", ""):
            data = ""
        if proj not in linhas_por_proj:
            linhas_por_proj[proj] = []
        linhas_por_proj[proj].append({"tipo": tipo, "arquivo": arq, "valor": valor, "data": data})

    dados       = ler_db()
    atualizados = []
    criados     = []

    for proj_num, itens in linhas_por_proj.items():
        idx = next((i for i, d in enumerate(dados)
                    if str(d.get("projeto", "")).upper() == proj_num), -1)

        # Cria registro novo se projeto não existir no banco
        if idx == -1:
            novo = {
                "projeto": proj_num, "ae": "", "aes": [],
                "us": "", "us_num": 0.0, "prazo": "", "prazo_alerta": "",
                "local": "", "custo": "", "status": "Pendente", "executado": False,
                "arquivo_ae": "", "pasta_projeto": "", "pdfs_projeto": [],
                "bmds": [], "ffos": [], "medicao1": "", "medicao2": "",
                "final": "", "origem": ["importado_excel"], "avisos": [],
            }
            dados.append(novo)
            idx = len(dados) - 1
            criados.append(proj_num)

        r    = dados[idx]
        arqs_bmds = {b.get("arquivo", "") for b in r.get("bmds", [])}
        arqs_ffos = {f.get("arquivo", "") for f in r.get("ffos", [])}
        valor_total_proj = 0.0

        for item in itens:
            valor_total_proj += item["valor"]
            entrada = {
                "arquivo":      item["arquivo"],
                "tipo":         item["tipo"],
                "data":         item["data"],
                "rateio":       _fmt_brl(item["valor"]).replace("R$ ", "") if item["valor"] > 0 else "",
                "rateio_float": item["valor"],
                "origem":       "importado_excel",
            }
            if item["tipo"] == "FFO":
                if item["arquivo"] not in arqs_ffos:
                    r.setdefault("ffos", []).append(entrada)
                    arqs_ffos.add(item["arquivo"])
                else:
                    # Atualiza data/rateio no registro existente
                    for f in r.get("ffos", []):
                        if f.get("arquivo") == item["arquivo"]:
                            if item["data"] and not f.get("data"):
                                f["data"] = item["data"]
                            if item["valor"] > 0 and not f.get("rateio_float"):
                                f["rateio"] = entrada["rateio"]
                                f["rateio_float"] = item["valor"]
                            break
                if not r.get("final"):
                    data_ffo = item["data"] or ""
                    r["final"] = data_ffo or item["arquivo"]
                    r["executado"] = True
                    r["status"] = "Final"
            else:
                if item["arquivo"] not in arqs_bmds:
                    r.setdefault("bmds", []).append(entrada)
                    arqs_bmds.add(item["arquivo"])
                else:
                    # Atualiza data/rateio no registro existente
                    for b in r.get("bmds", []):
                        if b.get("arquivo") == item["arquivo"]:
                            if item["data"] and not b.get("data"):
                                b["data"] = item["data"]
                            if item["valor"] > 0 and not b.get("rateio_float"):
                                b["rateio"] = entrada["rateio"]
                                b["rateio_float"] = item["valor"]
                            break
                bmds_ord = sorted(r["bmds"], key=lambda x: x.get("data", "") or "")
                if len(bmds_ord) >= 1 and not r.get("medicao1"):
                    r["medicao1"] = bmds_ord[0].get("data") or bmds_ord[0].get("arquivo", "")
                if len(bmds_ord) >= 2 and not r.get("medicao2"):
                    r["medicao2"] = bmds_ord[1].get("data") or bmds_ord[1].get("arquivo", "")

        dados[idx] = r
        atualizados.append({"projeto": proj_num, "valor": _fmt_brl(valor_total_proj), "criado": proj_num in criados})

    salvar_db(dados)
    return jsonify({
        "ok":               True,
        "total_importados": len(linhas_por_proj),
        "atualizados":      len(atualizados),
        "criados":          len(criados),
        "projetos_criados": criados,
        "detalhes":         atualizados,
    })


# ─────────────────────────────────────────────────────────────────────
#  API: listar PDFs
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/pdfs_pasta")
def api_pdfs_pasta():
    pasta   = request.args.get("pasta", "").strip()
    IGNORAR  = {"thumbs.db", ".ds_store", "desktop.ini"}
    arquivos = []
    # Tenta pasta local
    caminho = os.path.join(PASTA_PROJETOS, pasta)
    if os.path.isdir(caminho):
        for raiz, dirs, files in os.walk(caminho):
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for f in sorted(files):
                if f.lower() in IGNORAR or f.startswith('.'):
                    continue
                caminho_rel = os.path.relpath(os.path.join(raiz, f), caminho)
                arquivos.append(caminho_rel.replace("\\", "/"))
        if arquivos:
            return jsonify(sorted(arquivos))
    # Fallback: PostgreSQL
    if DATABASE_URL:
        pg_arquivos = _listar_arquivos_pg('projetos')
        if pasta:
            # Arquivos salvos como 'pasta/arquivo.pdf' ou 'pasta/sub/arquivo.pdf'
            prefixo = pasta + '/'
            filtrados = [n for n in pg_arquivos if n.startswith(prefixo)]
            if filtrados:
                # Retorna o caminho relativo DENTRO da pasta (sem o prefixo)
                return jsonify(sorted([n[len(prefixo):] for n in filtrados]))
        if pg_arquivos:
            return jsonify(sorted(pg_arquivos))
    return jsonify([])


@app.route("/api/pdfs_medicoes")
def api_pdfs_medicoes():
    projeto = request.args.get("projeto", "").strip()
    dados   = ler_db()
    idx     = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"bmds": [], "ffos": []})
    r = dados[idx]
    return jsonify({"bmds": r.get("bmds", []), "ffos": r.get("ffos", [])})


# ─────────────────────────────────────────────────────────────────────
#  Servir arquivos
# ─────────────────────────────────────────────────────────────────────

def _servir_arquivo(pasta_local, pasta_pg, filename, pg_chaves_extras=None):
    """
    Tenta servir arquivo do disco local; se não encontrar, busca no PostgreSQL.
    pg_chaves_extras: lista adicional de chaves para tentar no PG (ex: 'pasta/arquivo.pdf')
    """
    from flask import Response
    # 1) Tenta disco local (busca também recursivamente)
    caminho_local = os.path.join(pasta_local, filename)
    if os.path.exists(caminho_local):
        return send_from_directory(pasta_local, filename)
    # Busca recursiva no disco local
    nome_arq = filename.split('/')[-1]
    for raiz, dirs, files in os.walk(pasta_local):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        if nome_arq in files:
            return send_from_directory(raiz, nome_arq)
    # 2) Fallback: PostgreSQL
    if DATABASE_URL:
        tentativas = list(dict.fromkeys(
            [nome_arq, filename] + (pg_chaves_extras or [])
        ))
        for tentativa in tentativas:
            try:
                conteudo, mimetype = _ler_arquivo_pg(pasta_pg, tentativa)
                if conteudo is not None:
                    print(f'[Serve] ✅ Servindo {pasta_pg}/{tentativa} do PostgreSQL')
                    resp = Response(conteudo, mimetype=mimetype or 'application/octet-stream')
                    safe_name = nome_arq.replace('"', '')
                    disp = 'inline' if (mimetype or '').startswith(('application/pdf', 'image/')) else 'attachment'
                    resp.headers['Content-Disposition'] = f'{disp}; filename="{safe_name}"'
                    resp.headers['Content-Length'] = len(conteudo)
                    resp.headers['Access-Control-Allow-Origin'] = '*'
                    return resp
            except Exception as e:
                print(f'[Serve] Erro ao buscar {pasta_pg}/{tentativa}: {e}')
        # 3) Busca em TODAS as pastas do PG pelo nome do arquivo (último recurso)
        try:
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        'SELECT pasta, conteudo, mimetype FROM hagap_files WHERE nome=%s OR nome LIKE %s LIMIT 1',
                        (nome_arq, f'%/{nome_arq}')
                    )
                    row = cur.fetchone()
                    if row:
                        conteudo = bytes(row[1]) if not isinstance(row[1], bytes) else row[1]
                        mimetype = row[2] or 'application/octet-stream'
                        print(f'[Serve] ✅ Servindo {row[0]}/{nome_arq} do PG (busca global)')
                        resp = Response(conteudo, mimetype=mimetype)
                        safe_name = nome_arq.replace('"', '')
                        disp = 'inline' if mimetype.startswith(('application/pdf', 'image/')) else 'attachment'
                        resp.headers['Content-Disposition'] = f'{disp}; filename="{safe_name}"'
                        resp.headers['Content-Length'] = len(conteudo)
                        resp.headers['Access-Control-Allow-Origin'] = '*'
                        return resp
        except Exception as e:
            print(f'[Serve] Erro na busca global PG: {e}')
        print(f'[Serve] ❌ Não encontrado: pasta={pasta_pg}, arquivo={filename}')
    return jsonify({'erro': f'Arquivo não encontrado: {filename}', 'pasta': pasta_pg,
                    'dica': f'Acesse /api/debug_arquivo?pasta={pasta_pg}&nome={filename} para diagnóstico'}), 404


@app.route("/pdfs/aes/<path:filename>")
def servir_ae(filename):
    return _servir_arquivo(PASTA_AES, 'aes', filename)

@app.route("/pdfs/projetos/<pasta>/<path:filename>")
def servir_projeto(pasta, filename):
    # Gera chaves extras para o PG: o arquivo pode estar salvo como 'pasta/arquivo.pdf'
    pg_extras = [
        f"{pasta}/{filename}",
        f"{pasta}/{filename.split('/')[-1]}",
    ]
    return _servir_arquivo(os.path.join(PASTA_PROJETOS, pasta), 'projetos', filename, pg_extras)

@app.route("/pdfs/medicoes/<path:filename>")
def servir_medicao(filename):
    return _servir_arquivo(PASTA_MEDICOES, 'medicoes', filename)


# ─────────────────────────────────────────────────────────────────────
#  API: exportar Excel
# ─────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────
#  API: relação BMD vs AE (todos os projetos com medições)
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/relacao_bmd_ae")
def api_relacao_bmd_ae():
    dados = ler_db()
    resultado = []
    total_bmd = 0.0
    total_ae  = 0.0
    for d in dados:
        bmds = d.get("bmds", [])
        ffos = d.get("ffos", [])
        if not bmds and not ffos:
            continue
        soma_bmd = sum((b.get("rateio_float") or 0) for b in bmds)
        soma_ffo = sum((f.get("rateio_float") or 0) for f in ffos)
        soma_med = soma_bmd + soma_ffo
        custo_ae = _to_float(d.get("custo", ""))
        diff     = soma_med - custo_ae
        pct      = round(diff / custo_ae * 100, 2) if custo_ae > 0 else None
        situacao = "lucro" if diff > 0 else ("prejuizo" if diff < 0 else "neutro")
        total_bmd += soma_med
        total_ae  += custo_ae
        resultado.append({
            "projeto":       d.get("projeto", ""),
            "ae":            d.get("ae", ""),
            "local":         d.get("local", ""),
            "us":            d.get("us", ""),
            "status":        d.get("status", ""),
            "custo_ae":      custo_ae,
            "custo_ae_fmt":  _fmt_brl(custo_ae) if custo_ae else "—",
            "total_bmd":     soma_bmd,
            "total_bmd_fmt": _fmt_brl(soma_bmd) if soma_bmd else "—",
            "total_ffo":     soma_ffo,
            "total_med":     soma_med,
            "total_med_fmt": _fmt_brl(soma_med) if soma_med else "—",
            "diferenca":     diff,
            "diferenca_fmt": _fmt_brl(diff),
            "pct":           pct,
            "situacao":      situacao,
            "qtd_bmds":      len(bmds),
            "qtd_ffos":      len(ffos),
        })
    resultado.sort(key=lambda x: x["total_med"], reverse=True)
    diff_total = total_bmd - total_ae
    pct_total  = round(diff_total / total_ae * 100, 2) if total_ae > 0 else None
    return jsonify({
        "projetos":        resultado,
        "total_bmd":       total_bmd,
        "total_bmd_fmt":   _fmt_brl(total_bmd),
        "total_ae":        total_ae,
        "total_ae_fmt":    _fmt_brl(total_ae),
        "diferenca_total": diff_total,
        "diferenca_fmt":   _fmt_brl(diff_total),
        "pct_total":       pct_total,
        "situacao_total":  "lucro" if diff_total > 0 else ("prejuizo" if diff_total < 0 else "neutro"),
        "qtd_com_bmd":     len(resultado),
    })


@app.route("/api/exportar")
def api_exportar():
    dados  = ler_db()
    linhas = []
    for d in dados:
        row = {k: v for k, v in d.items()
               if k not in ("bmds", "ffos", "pdfs_projeto", "avisos", "origem", "aes")}
        row["qtd_bmds"] = len(d.get("bmds", []))
        row["qtd_ffos"] = len(d.get("ffos", []))
        row["qtd_aes"]  = len(d.get("aes", []))
        row["qtd_pdfs"] = len(d.get("pdfs_projeto", []))
        row["avisos"]   = " | ".join(d.get("avisos", []))
        row["origem"]   = ", ".join(d.get("origem", []))
        linhas.append(row)
    df = pd.DataFrame(linhas)
    df.to_excel(EXCEL_PATH, index=False)
    return send_from_directory(".", EXCEL_PATH, as_attachment=True,
                               download_name="HAGAP_resultado.xlsx")


# ─────────────────────────────────────────────────────────────────────
#  API: upload PDF
# ─────────────────────────────────────────────────────────────────────

@app.route('/api/upload_pdf', methods=['POST'])
def upload_pdf():
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'arquivo nao enviado'}), 400
    arq = request.files['arquivo']
    if arq.filename == '':
        return jsonify({'erro': 'arquivo vazio'}), 400
    # Sem restrição de tipo — aceita qualquer arquivo
    nome      = arq.filename
    conteudo  = arq.read()
    mimetype  = arq.mimetype or 'application/octet-stream'
    pasta_key     = request.form.get('pasta', 'uploads')  # 'aes', 'medicoes', 'projetos', 'uploads'
    projeto_nome  = request.form.get('projeto', '').strip()
    caminho_rel   = request.form.get('caminho_relativo', '').strip()  # ex: "1704877/subfolder/arquivo.pdf"
    skip_drive    = request.form.get('skip_drive', '0') == '1'  # pula upload Drive (sincronização em massa)
    destinos  = {'uploads': 'uploads', 'aes': PASTA_AES,
                 'projetos': PASTA_PROJETOS, 'medicoes': PASTA_MEDICOES}
    destino   = destinos.get(pasta_key, 'uploads')

    # Salva sempre no filesystem local (funciona tanto localmente quanto em modo híbrido)
    os.makedirs(destino, exist_ok=True)
    try:
        with open(os.path.join(destino, nome), 'wb') as f:
            f.write(conteudo)
    except Exception as e:
        print(f'[Upload] Aviso ao salvar no disco: {e}')

    # Salva no PostgreSQL quando rodando no Render (garante persistência entre restarts)
    pg_saved = False
    if DATABASE_URL:
        pg_pasta = pasta_key  # 'aes', 'medicoes', 'projetos', 'uploads'
        # Para arquivos de projeto, salva com prefixo do projeto para evitar colissões
        # Formato: nome_pg = "<projeto>/<subpasta_opcional>/<arquivo>" ou só "<arquivo>"
        if pasta_key == 'projetos' and projeto_nome:
            if caminho_rel:
                # Remove o primeiro segmento do caminho (que é a pasta raíz selecionada)
                partes = caminho_rel.replace('\\', '/').split('/')
                sub = '/'.join(partes[1:]) if len(partes) > 1 else partes[-1]
                nome_pg = f"{projeto_nome}/{sub}"
            else:
                nome_pg = f"{projeto_nome}/{nome}"
        else:
            nome_pg = nome
        pg_saved = _salvar_arquivo_pg(pg_pasta, nome_pg, conteudo, mimetype)
        if pg_saved:
            print(f'[Files] ✅ {nome_pg} salvo no PostgreSQL (pasta={pg_pasta})')
            # Salva também com o nome simples (sem prefixo de projeto) para facilitar lookup
            if nome_pg != nome:
                _salvar_arquivo_pg(pg_pasta, nome, conteudo, mimetype)
                print(f'[Files] ✅ {nome} salvo também como nome simples (pasta={pg_pasta})')

    # Tenta Google Drive como complemento (pulado quando skip_drive=1)
    drive_url = None
    service   = None if skip_drive else _get_drive_service()
    if service:
        folder_id = _find_drive_folder(service)
        if folder_id:
            from googleapiclient.http import MediaIoBaseUpload
            try:
                uploaded = service.files().create(
                    body={"name": nome, "parents": [folder_id]},
                    media_body=MediaIoBaseUpload(io.BytesIO(conteudo), mimetype=mimetype, resumable=True),
                    fields="id, webViewLink"
                ).execute()
                drive_url = uploaded.get("webViewLink")
                try:
                    service.permissions().create(
                        fileId=uploaded["id"],
                        body={"type": "anyone", "role": "reader"}
                    ).execute()
                except Exception:
                    pass
            except Exception as e:
                print(f"[Drive] Erro: {e}")

    result = {'ok': True, 'arquivo': nome, 'pasta': destino, 'pg': pg_saved}
    if drive_url:
        result['drive'] = True
        result['url']   = drive_url
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────
#  API: deletar arquivo individual
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/deletar_arquivo", methods=["POST"])
def api_deletar_arquivo():
    """Remove um arquivo vinculado de um projeto (BMD, FFO, AE ou extra)."""
    body    = request.get_json() or {}
    projeto = body.get("projeto", "").strip()
    arquivo = body.get("arquivo", "").strip()
    tipo    = body.get("tipo", "").lower()   # 'bmd', 'ffo', 'ae', 'outro'
    if not projeto or not arquivo:
        return jsonify({"erro": "Campos 'projeto' e 'arquivo' obrigatórios"}), 400
    dados = ler_db()
    idx = encontrar_idx(dados, projeto)
    if idx == -1:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    r = dados[idx]
    removido = False
    if tipo == "bmd":
        antes = len(r.get("bmds", []))
        r["bmds"] = [b for b in r.get("bmds", []) if b.get("arquivo") != arquivo]
        removido = len(r["bmds"]) < antes
        bmds_ord = sorted(r["bmds"], key=lambda x: x.get("data", "") or "")
        r["medicao1"] = bmds_ord[0].get("data") or bmds_ord[0].get("arquivo", "") if len(bmds_ord) >= 1 else ""
        r["medicao2"] = bmds_ord[1].get("data") or bmds_ord[1].get("arquivo", "") if len(bmds_ord) >= 2 else ""
    elif tipo == "ffo":
        antes = len(r.get("ffos", []))
        r["ffos"] = [f for f in r.get("ffos", []) if f.get("arquivo") != arquivo]
        removido = len(r["ffos"]) < antes
        if not r["ffos"]:
            r["executado"] = False
            r["status"] = "Pendente" if not r.get("bmds") else r.get("status", "Pendente")
            r["final"] = ""
        else:
            r["final"] = r["ffos"][-1].get("data") or r["ffos"][-1].get("arquivo", "")
    elif tipo == "ae":
        antes = len(r.get("aes", []))
        r["aes"] = [a for a in r.get("aes", []) if a.get("arquivo") != arquivo]
        removido = len(r["aes"]) < antes
        if r.get("arquivo_ae") == arquivo:
            r["arquivo_ae"] = r["aes"][0]["arquivo"] if r["aes"] else ""
    else:
        antes = len(r.get("arquivos_extras", []))
        r["arquivos_extras"] = [a for a in r.get("arquivos_extras", []) if a.get("arquivo") != arquivo]
        removido = len(r.get("arquivos_extras", [])) < antes
    # Tenta deletar do PostgreSQL também
    if DATABASE_URL and removido:
        pasta_map = {"bmd": "medicoes", "ffo": "medicoes", "ae": "aes", "outro": "projetos"}
        pasta_pg  = pasta_map.get(tipo, "uploads")
        try:
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM hagap_files WHERE pasta=%s AND nome=%s", (pasta_pg, arquivo))
                conn.commit()
        except Exception as e:
            print(f"[Delete] Erro ao deletar do PG: {e}")
    salvar_db(dados)
    return jsonify({"ok": True, "removido": removido, "registro": r})


@app.route("/api/deletar_arquivo_pg", methods=["POST"])
def api_deletar_arquivo_pg():
    """Deleta um arquivo diretamente do PostgreSQL (sem vínculo com projeto no DB)."""
    body  = request.get_json() or {}
    pasta = body.get("pasta", "").strip()   # 'aes', 'medicoes', 'projetos', 'uploads'
    nome  = body.get("nome",  "").strip()   # pode conter subpasta: 'projeto/arquivo.pdf'
    if not pasta or not nome:
        return jsonify({"erro": "Campos 'pasta' e 'nome' obrigatórios"}), 400
    if not DATABASE_URL:
        return jsonify({"erro": "Apenas disponível no modo Render"}), 400
    try:
        _ensure_pg_files()
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM hagap_files WHERE pasta=%s AND nome=%s", (pasta, nome))
                deletados = cur.rowcount
            conn.commit()
        return jsonify({"ok": True, "deletados": deletados})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route('/uploads/<path:nome>')
def uploads(nome):
    # Tenta disco local primeiro
    caminho = os.path.join('uploads', nome)
    if os.path.exists(caminho):
        return send_from_directory('uploads', nome)
    # Fallback: busca no PostgreSQL (pasta uploads ou busca global)
    if DATABASE_URL:
        from flask import Response
        nome_base = nome.split('/')[-1]
        for tentativa in [nome_base, nome]:
            conteudo, mimetype = _ler_arquivo_pg('uploads', tentativa)
            if conteudo is not None:
                resp = Response(conteudo, mimetype=mimetype or 'application/octet-stream')
                resp.headers['Content-Disposition'] = f'inline; filename="{nome_base}"'
                resp.headers['Content-Length'] = len(conteudo)
                return resp
        # Busca global no PG
        try:
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        'SELECT pasta, conteudo, mimetype FROM hagap_files WHERE nome=%s OR nome LIKE %s LIMIT 1',
                        (nome_base, f'%/{nome_base}')
                    )
                    row = cur.fetchone()
                    if row:
                        conteudo = bytes(row[1]) if not isinstance(row[1], bytes) else row[1]
                        mimetype = row[2] or 'application/octet-stream'
                        resp = Response(conteudo, mimetype=mimetype)
                        resp.headers['Content-Disposition'] = f'inline; filename="{nome_base}"'
                        resp.headers['Content-Length'] = len(conteudo)
                        return resp
        except Exception:
            pass
    return jsonify({'erro': f'Arquivo não encontrado: {nome}'}), 404


@app.route('/api/info_pastas')
def api_info_pastas():
    def _abs(p):
        return p if os.path.isabs(p) else os.path.join(os.path.abspath('.'), p)
    return jsonify({
        'modo':            'render' if DATABASE_URL else 'local',
        'aes':             _abs(PASTA_AES),
        'projetos':        _abs(PASTA_PROJETOS),
        'medicoes':        _abs(PASTA_MEDICOES),
        'uploads':         os.path.join(os.path.abspath('.'), 'uploads'),
        'aes_existe':      os.path.isdir(PASTA_AES),
        'projetos_existe': os.path.isdir(PASTA_PROJETOS),
        'medicoes_existe': os.path.isdir(PASTA_MEDICOES),
    })


# ─────────────────────────────────────────────────────────────────────
#  API: debug arquivos (diagnóstico de 404)
# ─────────────────────────────────────────────────────────────────────

@app.route("/api/debug_arquivo")
def api_debug_arquivo():
    """Diagnóstico: verifica se um arquivo existe no disco e/ou no PostgreSQL."""
    pasta_pg  = request.args.get("pasta", "").strip()    # 'aes', 'medicoes', 'projetos'
    nome      = request.args.get("nome",  "").strip()
    info = {
        "pasta_pg":   pasta_pg,
        "nome":       nome,
        "DATABASE_URL_set": bool(DATABASE_URL),
        "disco":      {},
        "postgres":   {},
    }
    # Verifica disco
    mapa_pasta = {"aes": PASTA_AES, "medicoes": PASTA_MEDICOES, "projetos": PASTA_PROJETOS, "uploads": "uploads"}
    pasta_local = mapa_pasta.get(pasta_pg, pasta_pg)
    caminho_local = os.path.join(pasta_local, nome) if nome else pasta_local
    info["disco"]["pasta_local"]    = pasta_local
    info["disco"]["caminho_local"]  = caminho_local
    info["disco"]["pasta_existe"]   = os.path.isdir(pasta_local)
    info["disco"]["arquivo_existe"] = os.path.exists(caminho_local) if nome else False
    # Lista arquivos na pasta local
    if os.path.isdir(pasta_local):
        try:
            todos = []
            for raiz, _, files in os.walk(pasta_local):
                for f in files:
                    todos.append(os.path.relpath(os.path.join(raiz, f), pasta_local).replace("\\", "/"))
            info["disco"]["arquivos"] = sorted(todos)[:50]
        except Exception as e:
            info["disco"]["erro"] = str(e)
    # Verifica PostgreSQL
    if DATABASE_URL:
        try:
            _ensure_pg_files()
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    # Verifica tabela
                    cur.execute("SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_name='hagap_files')")
                    info["postgres"]["tabela_existe"] = cur.fetchone()[0]
                    if info["postgres"]["tabela_existe"]:
                        # Total de arquivos no banco
                        cur.execute("SELECT pasta, COUNT(*) FROM hagap_files GROUP BY pasta ORDER BY pasta")
                        info["postgres"]["contagem_por_pasta"] = {r[0]: r[1] for r in cur.fetchall()}
                        # Lista arquivos da pasta solicitada
                        if pasta_pg:
                            cur.execute("SELECT nome FROM hagap_files WHERE pasta=%s ORDER BY nome", (pasta_pg,))
                            info["postgres"]["arquivos_na_pasta"] = [r[0] for r in cur.fetchall()]
                        # Verifica arquivo específico
                        if nome and pasta_pg:
                            cur.execute("SELECT nome, mimetype, length(conteudo) FROM hagap_files WHERE pasta=%s AND nome=%s", (pasta_pg, nome))
                            row = cur.fetchone()
                            info["postgres"]["arquivo_encontrado"] = bool(row)
                            if row:
                                info["postgres"]["tamanho_bytes"] = row[2]
                                info["postgres"]["mimetype"] = row[1]
        except Exception as e:
            info["postgres"]["erro"] = str(e)
    return jsonify(info)


@app.route("/api/listar_arquivos_pg")
def api_listar_arquivos_pg():
    """Lista todos os arquivos salvos no PostgreSQL."""
    if not DATABASE_URL:
        return jsonify({"erro": "Apenas disponível no modo Render (DATABASE_URL)"}), 400
    try:
        _ensure_pg_files()
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT pasta, nome, mimetype, length(conteudo), criado_at FROM hagap_files ORDER BY pasta, nome")
                rows = cur.fetchall()
        return jsonify([{"pasta": r[0], "nome": r[1], "mimetype": r[2],
                         "tamanho_bytes": r[3], "criado_at": str(r[4])} for r in rows])
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/avisos")
def api_avisos():
    dados  = ler_db()
    avisos = []
    for d in dados:
        for av in d.get("avisos", []):
            avisos.append({"projeto": d["projeto"], "aviso": av})
        if not d.get("ae") and not d.get("aes"):
            avisos.append({"projeto": d["projeto"], "aviso": "Sem AE vinculada"})
        if not d.get("local"):
            avisos.append({"projeto": d["projeto"], "aviso": "Local não encontrado"})
        if not d.get("prazo"):
            avisos.append({"projeto": d["projeto"], "aviso": "Prazo não encontrado"})
    return jsonify(avisos)


@app.route("/api/extrair_rateio", methods=["POST"])
def api_extrair_rateio():
    body    = request.get_json() or {}
    arquivo = body.get("arquivo", "").strip()
    custo   = body.get("custo", "0")
    caminho = os.path.join(PASTA_MEDICOES, arquivo)
    if not os.path.exists(caminho):
        return jsonify({"erro": f"Arquivo não encontrado: {arquivo}"}), 404
    texto        = _extrair_texto_medicao(caminho)
    rateio_str, us_str = _extrair_rateio_e_us(texto)
    rateio_float = _to_float(rateio_str)
    custo_float  = _to_float(custo)
    pct = round(rateio_float / custo_float * 100, 2) if custo_float > 0 and rateio_float > 0 else None
    return jsonify({"rateio_str": rateio_str, "rateio": rateio_float,
                    "us": us_str, "porcentagem": pct})


@app.route("/api/limpar_nomes_projetos", methods=["POST"])
def api_limpar_nomes_projetos():
    """Remove espaços, \\r, \\n e \\t dos nomes de projetos no banco de dados."""
    dados = ler_db()
    corrigidos = []
    vistos = set()
    novos_dados = []
    for d in dados:
        nome_original = d.get('projeto', '')
        nome_limpo = re.sub(r'[\r\n\t]', '', nome_original.strip())
        if nome_limpo != nome_original:
            corrigidos.append({'de': repr(nome_original), 'para': nome_limpo})
            d['projeto'] = nome_limpo
        # Remove duplicatas geradas pela limpeza
        chave = nome_limpo.upper()
        if chave in vistos:
            corrigidos.append({'de': nome_limpo, 'para': 'DUPLICATA REMOVIDA'})
            continue
        vistos.add(chave)
        novos_dados.append(d)
    salvar_db(novos_dados)
    return jsonify({'ok': True, 'total': len(dados), 'corrigidos': len(corrigidos), 'detalhes': corrigidos})


@app.route("/api/reprocessar_bmds", methods=["POST"])
def api_reprocessar_bmds():
    dados      = ler_db()
    atualizados = 0
    for row in dados:
        for bmd in row.get("bmds", []):
            if bmd.get("rateio_float", 0) > 0:
                continue
            arquivo = bmd.get("arquivo", "")
            if not arquivo:
                continue
            cam = os.path.join(PASTA_MEDICOES, arquivo)
            if os.path.exists(cam):
                texto = _extrair_texto_medicao(cam)
                rateio_str, us_med = _extrair_rateio_e_us(texto)
                if rateio_str:
                    bmd["rateio"]       = rateio_str
                    bmd["rateio_float"] = _to_float(rateio_str)
                    atualizados += 1
                if us_med and not bmd.get("us"):
                    bmd["us"] = us_med
    salvar_db(dados)
    return jsonify({"ok": True, "atualizados": atualizados})



@app.route("/api/listar_todos_arquivos")
def api_listar_todos_arquivos():
    """Lista todos os arquivos em todas as pastas (disco + PostgreSQL)."""
    pastas = {'aes': PASTA_AES, 'medicoes': PASTA_MEDICOES, 'projetos': PASTA_PROJETOS, 'uploads': 'uploads'}
    resultado = {}
    for pasta_key, pasta_local in pastas.items():
        arquivos = []
        # Disco local
        if os.path.isdir(pasta_local):
            for raiz, dirs, files in os.walk(pasta_local):
                dirs[:] = [d for d in dirs if not d.startswith('.')]
                for f in sorted(files):
                    if f.startswith('.'): continue
                    rel = os.path.relpath(os.path.join(raiz, f), pasta_local).replace('\\', '/')
                    arquivos.append({'nome': rel, 'origem': 'disco'})
        # PostgreSQL
        if DATABASE_URL:
            pg_list = _listar_arquivos_pg(pasta_key)
            pg_nomes = {a['nome'] for a in arquivos}
            for nome in pg_list:
                if nome not in pg_nomes:
                    arquivos.append({'nome': nome, 'origem': 'pg'})
        resultado[pasta_key] = sorted(arquivos, key=lambda x: x['nome'])
    return jsonify(resultado)


# ─────────────────────────────────────────────────────────────────────
#  API: materiais BMD/FFO por projeto (lê materiais_por_projeto.xlsx)
# ─────────────────────────────────────────────────────────────────────

MATERIAIS_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "materiais_por_projeto.xlsx")

@app.route("/api/materiais_bmd")
def api_materiais_bmd():
    """Lê materiais_por_projeto.xlsx e retorna dados das abas Resumo e Materiais_Detalhados."""
    # Tenta encontrar o arquivo
    caminhos = [
        MATERIAIS_XLSX,
        os.path.join(os.getcwd(), "materiais_por_projeto.xlsx"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "materiais_por_projeto.xlsx"),
    ]
    xlsx_path = next((c for c in caminhos if os.path.exists(c)), None)

    # Fallback: tenta buscar do PostgreSQL (caso tenha sido enviado via upload)
    if not xlsx_path and DATABASE_URL:
        try:
            conteudo, _ = _ler_arquivo_pg('uploads', 'materiais_por_projeto.xlsx')
            if conteudo:
                import tempfile
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmp.write(conteudo); tmp.close()
                xlsx_path = tmp.name
        except Exception:
            pass

    if not xlsx_path:
        return jsonify({"erro": "Arquivo materiais_por_projeto.xlsx não encontrado. Faça upload pelo botão de upload."}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return jsonify({"erro": f"Erro ao ler xlsx: {e}"}), 500

    resumo = []
    itens  = []

    # Aba Resumo
    if 'Resumo' in wb.sheetnames:
        ws = wb['Resumo']
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c else '' for c in row]
                continue
            if not any(row):
                continue
            r = dict(zip(header, row))
            resumo.append({
                'projeto':       str(r.get('projeto') or '').strip(),
                'arquivo':       str(r.get('arquivo') or '').strip(),
                'total_geral':   float(r.get('total_geral') or 0),
                'total_material': float(r.get('total_material') or 0),
                'total_mao_obra': float(r.get('total_mao_obra') or 0),
                'possui_material': str(r.get('possui_material') or 'NAO').strip(),
            })

    # Aba Materiais_Detalhados
    aba_det = next((s for s in wb.sheetnames if 'detalha' in s.lower()), None)
    if aba_det:
        ws = wb[aba_det]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c else '' for c in row]
                continue
            if not any(row):
                continue
            r = dict(zip(header, row))
            itens.append({
                'projeto':        str(r.get('projeto') or '').strip(),
                'arquivo':        str(r.get('arquivo') or '').strip(),
                'item':           str(r.get('item') or '').strip(),
                'codigo':         str(r.get('codigo') or '').strip(),
                'descricao':      str(r.get('descricao') or '').strip(),
                'quantidade':     float(r.get('quantidade') or 0),
                'valor_unitario': float(r.get('valor_unitario') or 0),
                'valor_total':    float(r.get('valor_total') or 0),
            })

    # Aplicar overrides do banco de dados
    overrides = {}
    checks = {}
    if DATABASE_URL:
        try:
            with _pg_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT projeto, arquivo, item_key, descricao, quantidade, valor_unitario, valor_total, is_deleted, is_checked, check_obs FROM hagap_mat_overrides")
                    for row in cur.fetchall():
                        k = (row[0], row[1], row[2])
                        overrides[k] = {
                            'descricao': row[3], 'quantidade': row[4],
                            'valor_unitario': row[5], 'valor_total': row[6],
                            'is_deleted': row[7]
                        }
                        checks[k] = {'is_checked': row[8], 'check_obs': row[9] or ''}
        except Exception as e:
            print(f'[mat_overrides] erro: {e}')

    # Aplica overrides nos itens
    itens_filtrados = []
    for it in itens:
        k = (it['projeto'], it['arquivo'], it.get('item','') + '|' + it.get('codigo',''))
        ov = overrides.get(k, {})
        if ov.get('is_deleted'):
            continue
        if ov:
            it = {**it}
            for f in ('descricao','quantidade','valor_unitario','valor_total'):
                if ov.get(f) is not None:
                    it[f] = ov[f]
        ck = checks.get(k, {})
        it['is_checked'] = ck.get('is_checked', False)
        it['check_obs']  = ck.get('check_obs', '')
        it['item_key']   = k[2]
        itens_filtrados.append(it)

    # Aplica checks nos resumo
    for rv in resumo:
        k_resumo = (rv['projeto'], rv['arquivo'], '__resumo__')
        ck = checks.get(k_resumo, {})
        rv['is_checked'] = ck.get('is_checked', False)
        rv['check_obs']  = ck.get('check_obs', '')

    return jsonify({'resumo': resumo, 'itens': itens_filtrados})


# ── Editar item de material ──────────────────────────────────────────
@app.route("/api/materiais_item_update", methods=["POST"])
def api_materiais_item_update():
    if not DATABASE_URL:
        return jsonify({"erro": "DATABASE_URL não configurado"}), 400
    body = request.get_json(force=True)
    projeto   = body.get('projeto', '')
    arquivo   = body.get('arquivo', '')
    item_key  = body.get('item_key', '')
    fields    = body.get('fields', {})
    if not projeto or not item_key:
        return jsonify({"erro": "projeto e item_key são obrigatórios"}), 400
    try:
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hagap_mat_overrides (projeto, arquivo, item_key, descricao, quantidade, valor_unitario, valor_total)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (projeto, arquivo, item_key) DO UPDATE SET
                        descricao      = COALESCE(EXCLUDED.descricao,      hagap_mat_overrides.descricao),
                        quantidade     = COALESCE(EXCLUDED.quantidade,     hagap_mat_overrides.quantidade),
                        valor_unitario = COALESCE(EXCLUDED.valor_unitario, hagap_mat_overrides.valor_unitario),
                        valor_total    = COALESCE(EXCLUDED.valor_total,    hagap_mat_overrides.valor_total),
                        updated_at     = NOW()
                """, (
                    projeto, arquivo, item_key,
                    fields.get('descricao'), fields.get('quantidade'),
                    fields.get('valor_unitario'), fields.get('valor_total')
                ))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ── Deletar item de material ─────────────────────────────────────────
@app.route("/api/materiais_item_delete", methods=["POST"])
def api_materiais_item_delete():
    if not DATABASE_URL:
        return jsonify({"erro": "DATABASE_URL não configurado"}), 400
    body = request.get_json(force=True)
    projeto  = body.get('projeto', '')
    arquivo  = body.get('arquivo', '')
    item_key = body.get('item_key', '')
    if not projeto or not item_key:
        return jsonify({"erro": "projeto e item_key são obrigatórios"}), 400
    try:
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hagap_mat_overrides (projeto, arquivo, item_key, is_deleted)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (projeto, arquivo, item_key) DO UPDATE SET is_deleted=TRUE, updated_at=NOW()
                """, (projeto, arquivo, item_key))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ── Checklist de item/resumo ─────────────────────────────────────────
@app.route("/api/materiais_item_check", methods=["POST"])
def api_materiais_item_check():
    if not DATABASE_URL:
        return jsonify({"erro": "DATABASE_URL não configurado"}), 400
    body = request.get_json(force=True)
    projeto   = body.get('projeto', '')
    arquivo   = body.get('arquivo', '')
    item_key  = body.get('item_key', '__resumo__')
    checked   = bool(body.get('checked', False))
    check_obs = body.get('check_obs', '')
    if not projeto:
        return jsonify({"erro": "projeto é obrigatório"}), 400
    try:
        with _pg_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO hagap_mat_overrides (projeto, arquivo, item_key, is_checked, check_obs)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (projeto, arquivo, item_key) DO UPDATE SET
                        is_checked = EXCLUDED.is_checked,
                        check_obs  = EXCLUDED.check_obs,
                        updated_at = NOW()
                """, (projeto, arquivo, item_key, checked, check_obs))
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ── Envio automático de BDO por e-mail ─────────────────────────────
@app.route("/api/enviar_bdo_email", methods=["POST"])
def api_enviar_bdo_email():
    """
    Recebe os dados do BDO e envia um e-mail formatado para
    gabrielmedeirosasp@gmail.com.

    Variáveis de ambiente necessárias no Render:
      EMAIL_REMETENTE  → seu e-mail Gmail remetente
      EMAIL_SENHA      → senha de app do Gmail (16 caracteres)
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText

    DESTINATARIO = "gabrielmedeirosasp@gmail.com"
    remetente    = os.environ.get("EMAIL_REMETENTE", "")
    senha        = os.environ.get("EMAIL_SENHA",     "")

    if not remetente or not senha:
        return jsonify({"ok": False, "erro": "EMAIL_REMETENTE ou EMAIL_SENHA não configurados"}), 200

    try:
        d = request.get_json(force=True) or {}

        srv      = ", ".join(d.get("servicos", [])) or "Nenhum"
        manobra  = d.get("manobra_texto", "—")
        alt      = d.get("alteracoes_texto", "—")
        omb_plv  = f"OMB: {d.get('num_omb','—')} | PLV: {d.get('num_plv','—')}" \
                   if d.get("servico_livre") is False else "Serviço Livre"

        corpo_html = f"""
        <html><body style="font-family:Arial,sans-serif;color:#0d1b3e;">
        <div style="max-width:620px;margin:0 auto;border:2px solid #b8c8e8;border-radius:10px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#061539,#0a2f7a);color:white;padding:16px 20px;">
            <h2 style="margin:0;font-size:16px;">📋 BDO — Boletim Diário de Obras</h2>
            <p style="margin:4px 0 0;font-size:11px;opacity:.75;">Hagap Instalações Elétricas</p>
          </div>
          <table style="width:100%;border-collapse:collapse;font-size:13px;">
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;width:40%;">Projeto</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{d.get('num_projeto','—')}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Cidade</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{d.get('cidade','—')}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Data</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{d.get('data_bdo','—')}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Horário</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{d.get('hora_inicio','—')} → {d.get('hora_final','—')}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Encarregado</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{d.get('encarregado','—')}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Serviço Livre</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{omb_plv}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Manobra</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{manobra}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Alterações</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{alt}</td></tr>
            <tr><td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;font-weight:700;color:#4a5a7a;">Serviços</td>
                <td style="padding:8px 14px;border-bottom:1px solid #e8ecf5;">{srv}</td></tr>
            {f'<tr><td style="padding:8px 14px;font-weight:700;color:#4a5a7a;">Observações</td><td style="padding:8px 14px;">{d.get("observacoes","")}</td></tr>' if d.get('observacoes') else ''}
          </table>
          <div style="background:#f0f4fb;padding:8px 14px;font-size:10px;color:#4a5a7a;text-align:center;">
            HAGAP Instalações Elétricas — Enviado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')}
          </div>
        </div></body></html>
        """

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"BDO — {d.get('num_projeto','Sem projeto')} | {d.get('data_bdo','')}"
        msg["From"]    = remetente
        msg["To"]      = DESTINATARIO
        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as smtp:
            smtp.login(remetente, senha)
            smtp.sendmail(remetente, DESTINATARIO, msg.as_string())

        return jsonify({"ok": True})

    except Exception as e:
        print(f"[E-mail BDO] Erro: {e}")
        return jsonify({"ok": False, "erro": str(e)}), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)